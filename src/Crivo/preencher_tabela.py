#!/usr/bin/env python3
"""
preencher_tabela.py
Lê resultados.txt gerado por rodar_crivo.sh e preenche Crivo.xlsx.

Uso:
    python preencher_tabela.py [resultados.txt] [Crivo.xlsx] [Crivo_preenchido.xlsx]
"""

import sys
import os
from openpyxl import load_workbook

# ---------- arquivos ---------------------------------------------------------
ARQ_RESULTADOS = sys.argv[1] if len(sys.argv) > 1 else "resultados.txt"
ARQ_ENTRADA    = sys.argv[2] if len(sys.argv) > 2 else "Crivo.xlsx"
ARQ_SAIDA      = sys.argv[3] if len(sys.argv) > 3 else "Crivo_preenchido.xlsx"

# ---------- mapeamento de colunas (N → coluna Excel) -------------------------
# Tempo e Speedup (colunas C..F e J..M)
COL_TEMPO_BOOL   = {100_000_000: "C", 1_000_000_000: "D",
                    10_000_000_000: "E", 100_000_000_000: "F"}
COL_TEMPO_BITSET = COL_TEMPO_BOOL          # mesmas colunas, linhas diferentes

# Memória e Primos (colunas J..M)
COL_DIR = {100_000_000: "J", 1_000_000_000: "K",
           10_000_000_000: "L", 100_000_000_000: "M"}

# Linhas de dados: rodada 1-10 → linha base + (rodada-1)
LINHA_BASE_TEMPO_BOOL   = 2   # rodada 1 → linha 2, rodada 10 → linha 11
LINHA_BASE_TEMPO_BITSET = 13  # rodada 1 → linha 13, rodada 10 → linha 22

LINHA_BASE_PRIMO_BOOL   = 5   # rodada 1 → linha 5, rodada 10 → linha 14
LINHA_BASE_PRIMO_BITSET = 16  # rodada 1 → linha 16, rodada 10 → linha 25

LINHA_BASE_MEM_BOOL   = 5    # mesmo layout de Primos(N)
LINHA_BASE_MEM_BITSET = 16

# Marcações: linha 44 e 45 (segunda tabela à direita, linhas 44-45)
# Conforme planilha: linha 45 tem "Marcações" em H45
# As marcações ficam em I44:L44 (BOOL) e I45:L45 (BITSET) — ajuste se necessário
LINHA_MARC_BOOL   = 44
LINHA_MARC_BITSET = 45

# ---------- leitura dos resultados -------------------------------------------
def ler_resultados(caminho):
    """Retorna dict: (programa_base, N, rodada) → {tempo, primos, marcacoes, mem_kb}"""
    dados = {}
    if not os.path.exists(caminho):
        print(f"ERRO: arquivo '{caminho}' não encontrado.")
        sys.exit(1)

    with open(caminho) as f:
        for linha in f:
            linha = linha.strip()
            if not linha or linha.startswith("#"):
                continue
            partes = linha.split()
            if len(partes) < 7:
                continue
            prog, N, rodada, tempo, primos, marcacoes, mem_kb = partes[:7]
            N = int(N); rodada = int(rodada)
            # normaliza nome do programa
            prog_base = "BOOL" if "BITSET" not in prog.upper() else "BITSET"
            dados[(prog_base, N, rodada)] = {
                "tempo":     float(tempo),
                "primos":    int(primos),
                "marcacoes": int(marcacoes),
                "mem_kb":    float(mem_kb) / 1024.0,  # converte KB → MB
            }
    return dados

# ---------- preenchimento ----------------------------------------------------
def preencher(dados, arq_in, arq_out):
    wb = load_workbook(arq_in)
    ws = wb.active

    VALORES_N = [100_000_000, 1_000_000_000, 10_000_000_000, 100_000_000_000]

    for N in VALORES_N:
        for rodada in range(1, 11):
            for prog in ("BOOL", "BITSET"):
                chave = (prog, N, rodada)
                if chave not in dados:
                    continue
                d = dados[chave]

                # ---- Tempo de Execução (segundos) ----
                col_t = COL_TEMPO_BOOL[N]  # mesma coluna para ambos
                if prog == "BOOL":
                    linha_t = LINHA_BASE_TEMPO_BOOL + (rodada - 1)
                else:
                    linha_t = LINHA_BASE_TEMPO_BITSET + (rodada - 1)
                ws[f"{col_t}{linha_t}"] = round(d["tempo"], 6)

                # ---- Primos(N) ----
                col_d = COL_DIR[N]
                if prog == "BOOL":
                    linha_p = LINHA_BASE_PRIMO_BOOL + (rodada - 1)
                else:
                    linha_p = LINHA_BASE_PRIMO_BITSET + (rodada - 1)
                ws[f"{col_d}{linha_p}"] = d["primos"]

                # ---- Memória (MB) ----
                if prog == "BOOL":
                    linha_m = LINHA_BASE_MEM_BOOL + (rodada - 1)
                else:
                    linha_m = LINHA_BASE_MEM_BITSET + (rodada - 1)
                ws[f"{col_d}{linha_m}"] = round(d["mem_kb"], 3)

            # ---- Marcações (apenas uma por N, última rodada disponível) ----
            for prog, linha_marc in (("BOOL", LINHA_MARC_BOOL),
                                     ("BITSET", LINHA_MARC_BITSET)):
                chave = (prog, N, rodada)
                if chave in dados:
                    col_d = COL_DIR[N]
                    ws[f"{col_d}{linha_marc}"] = dados[chave]["marcacoes"]

    wb.save(arq_out)
    print(f"Planilha salva em: {arq_out}")

# ---------- main -------------------------------------------------------------
if __name__ == "__main__":
    print(f"Lendo resultados de: {ARQ_RESULTADOS}")
    dados = ler_resultados(ARQ_RESULTADOS)
    print(f"Registros lidos: {len(dados)}")
    print(f"Preenchendo: {ARQ_ENTRADA} → {ARQ_SAIDA}")
    preencher(dados, ARQ_ENTRADA, ARQ_SAIDA)
