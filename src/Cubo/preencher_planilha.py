#!/usr/bin/env python3
"""
preencher_planilha.py
=====================
Lê o arquivo resultados.csv gerado pelo rodar_experimentos.sh e preenche
automaticamente a planilha CuboProfundidade_1a20.xlsx.

Uso:
    python3 preencher_planilha.py [resultados.csv] [CuboProfundidade_1a20.xlsx]

Por padrão assume os nomes acima no diretório atual.
"""

import sys
import csv
import re
from pathlib import Path
from openpyxl import load_workbook

# -------------------------------------------------------------------
# Configurações
# -------------------------------------------------------------------
CSV_FILE  = sys.argv[1] if len(sys.argv) > 1 else "resultados.csv"
XLSX_FILE = sys.argv[2] if len(sys.argv) > 2 else "CuboProfundidade_1a20.xlsx"

THREADS_LIST = [4, 8, 16, 32]

# -------------------------------------------------------------------
# Mapeamento: (algo_nome, mov) → linha de início do bloco na planilha
# Construído dinamicamente lendo a aba "Resultados".
# -------------------------------------------------------------------

def build_block_map(ws):
    """Varre a aba Resultados e retorna dict {(algo, mov): first_data_row}."""
    block_map = {}
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        if v1 and v2 is not None and isinstance(v2, int):
            algo_str = str(v1).strip()
            if algo_str in ("Sequencial", "TBB – Fitness Paralelo", "TBB – Modelo de Ilhas"):
                block_map[(algo_str, int(v2))] = r
    return block_map


def col_for_threads(base_col, threads, threads_list=THREADS_LIST):
    """Dado a coluna base (1T/Seq) e um número de threads, retorna a coluna Excel."""
    if threads not in threads_list:
        raise ValueError(f"Threads {threads} não está em {threads_list}")
    return base_col + threads_list.index(threads) + 1


def fill_row(ws, data_row, algo, threads, tempo, fitness, taxa_sucesso, geracao):
    """Preenche as colunas corretas para uma linha de dados."""
    # Colunas de dados:
    # C4  = Tempo Sequencial         C9  = Fitness Sequencial
    # C5..C8 = Tempo TBB 4/8/16/32T  C10..C13 = Fitness TBB 4/8/16/32T
    # C28 = Taxa Sucesso Seq         C29..C32 = Taxa TBB threads (FitPar ou Ilhas)
    # C33 = Geração Seq              C34..C37 = Geração TBB threads

    if algo == "Sequencial":
        ws.cell(data_row, 4).value  = tempo
        ws.cell(data_row, 9).value  = fitness
        ws.cell(data_row, 28).value = taxa_sucesso
        ws.cell(data_row, 33).value = geracao
    else:
        # TBB – Fitness Paralelo ou TBB – Modelo de Ilhas
        tc = col_for_threads(4, threads)   # cols 5,6,7,8
        fc = col_for_threads(9, threads)   # cols 10,11,12,13
        sc = col_for_threads(28, threads)  # cols 29,30,31,32
        gc = col_for_threads(33, threads)  # cols 34,35,36,37
        ws.cell(data_row, tc).value = tempo
        ws.cell(data_row, fc).value = fitness
        ws.cell(data_row, sc).value = taxa_sucesso
        ws.cell(data_row, gc).value = geracao


# -------------------------------------------------------------------
# Main
# -------------------------------------------------------------------

def main():
    csv_path  = Path(CSV_FILE)
    xlsx_path = Path(XLSX_FILE)

    if not csv_path.exists():
        print(f"ERRO: arquivo CSV não encontrado: {csv_path}")
        sys.exit(1)
    if not xlsx_path.exists():
        print(f"ERRO: planilha não encontrada: {xlsx_path}")
        sys.exit(1)

    print(f"Carregando planilha: {xlsx_path}")
    wb = load_workbook(str(xlsx_path))
    ws = wb["Resultados"]

    print("Mapeando blocos na aba Resultados...")
    block_map = build_block_map(ws)
    print(f"  {len(block_map)} blocos encontrados (esperado: 60)")

    print(f"Lendo resultados: {csv_path}")
    erros = 0
    preenchidos = 0

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                algo       = row["Algoritmo"].strip()
                threads    = int(row["Threads"]) if row["Threads"].strip() not in ("", "1") else None
                mov        = int(row["Mov"])
                rep        = int(row["Repeticao"])
                tempo      = float(row["Tempo_s"])
                fitness    = float(row["Fitness"])
                resolvido  = row["Resolvido"].strip().upper()
                geracao    = int(row["Geracao"])
                taxa       = 100 if resolvido == "SIM" else 0
            except (KeyError, ValueError) as e:
                print(f"  AVISO: linha ignorada ({e}): {row}")
                erros += 1
                continue

            key = (algo, mov)
            if key not in block_map:
                print(f"  AVISO: bloco não encontrado para algo='{algo}' mov={mov}")
                erros += 1
                continue

            block_start = block_map[key]
            # rep vai de 1..10, linha de dados = block_start + rep - 1
            if rep < 1 or rep > 10:
                print(f"  AVISO: rep={rep} fora de 1..10, ignorado")
                erros += 1
                continue

            data_row = block_start + rep - 1

            t_val = threads if threads else None
            fill_row(ws, data_row, algo, t_val, tempo, fitness, taxa, geracao)
            preenchidos += 1

    print(f"\nDados preenchidos: {preenchidos} | Erros/avisos: {erros}")

    # Salva (sobrescreve o original)
    wb.save(str(xlsx_path))
    print(f"Planilha salva: {xlsx_path}")
    print("\nPronto! Abra a planilha para conferir as fórmulas automáticas.")


if __name__ == "__main__":
    main()
